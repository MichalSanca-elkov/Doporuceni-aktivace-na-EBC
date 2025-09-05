import pandas as pd
import os
import sys
import traceback
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from concurrent.futures import ProcessPoolExecutor, as_completed
import multiprocessing


def zpracuj_soubor(cesta, df_goods):
    filename = os.path.basename(cesta)
    print(f"\n🔄 Zpracovávám soubor: {filename}")
    try:
        wb_input = load_workbook(cesta, read_only=True)
        sheet_names_input = wb_input.sheetnames
        if not sheet_names_input:
            return f"⚠️ {filename}: žádné listy – pravděpodobně poškozený nebo neplatný soubor.", None
        df = pd.read_excel(cesta, sheet_name=sheet_names_input[0])
        df.columns = df.columns.str.strip()
    except InvalidFileException:
        return f"❌ {filename}: soubor nelze otevřít – není validní .xlsx", None
    except Exception as e:
        return f"⚠️ Nelze načíst soubor {filename}: {e}", None

    col_oid = df.columns[0]
    ebc_col = next((c for c in df.columns if "EBC požadovaný stav" in c), None)
    ebc2_col = next((c for c in df.columns if "EBC 2 požadovaný stav" in c), None)

    if not ebc_col or not ebc2_col:
        return f"⚠️ {filename} neobsahuje potřebné EBC sloupce, přeskočeno.", None

    df_filtered = df[
        (df[ebc_col].astype(str).str.lower() == "ano") |
        (df[ebc2_col].astype(str).str.lower() == "ano")
    ].copy()

    if df_filtered.empty:
        return f"ℹ️ {filename}: žádné položky s EBC = ano.", None

    for col in ['Fotografie', 'Anotace', 'OID_zbozi']:
        if col in df_filtered.columns:
            df_filtered.drop(columns=col, inplace=True)

    df_filtered = df_filtered.merge(
        df_goods[['OID_zbozi', 'Fotografie', 'Anotace']],
        left_on=col_oid,
        right_on='OID_zbozi',
        how='left'
    )

    foto_col = [col for col in df_filtered.columns if col.lower().startswith("fotografie")][-1]
    anotace_col = [col for col in df_filtered.columns if col.lower().startswith("anotace")][-1]

    def check_completeness(row):
        has_photo = pd.notna(row[foto_col]) and str(row[foto_col]).strip() != ""
        has_annotation = pd.notna(row[anotace_col]) and str(row[anotace_col]).strip() != ""
        if has_photo and has_annotation:
            return "Fotografie + Anotace"
        elif has_photo:
            return "Pouze fotografie"
        elif has_annotation:
            return "Pouze anotace"
        else:
            return "Chybí obojí"

    df_filtered["Stav dat"] = df_filtered.apply(check_completeness, axis=1)
    df_filtered["Zdrojový soubor"] = filename
    return f"✅ Zpracováno: {filename}", df_filtered


def main():
    # 🔁 Zjisti aktuální složku
    if getattr(sys, 'frozen', False):
        slozka = os.path.dirname(sys.executable)
    else:
        slozka = os.path.dirname(os.path.abspath(__file__))

    soubor_good = os.path.join(slozka, "good.xlsx")

    # --- Načtení good.xlsx ---
    try:
        wb = load_workbook(soubor_good, read_only=True)
        sheet_names = wb.sheetnames
        print(f"✅ Nalezené listy v good.xlsx: {sheet_names}")
        if not sheet_names:
            raise ValueError("Soubor good.xlsx neobsahuje žádné načitatelné listy.")
        df_goods = pd.read_excel(soubor_good, sheet_name=sheet_names[0])
        df_goods.columns = df_goods.columns.str.strip()
    except Exception as e:
        print(f"❌ Chyba při načítání good.xlsx: {e}")
        traceback.print_exc()
        input("\nStiskněte Enter pro ukončení...")
        sys.exit(1)

    # --- Sběr souborů ---
    vstupni_soubory = [
        os.path.join(slozka, f) for f in os.listdir(slozka)
        if f.endswith(".xlsx") and f.lower() != "good.xlsx"
    ]

    vysledky = []

    with ProcessPoolExecutor(max_workers=multiprocessing.cpu_count()) as executor:
        future_to_file = {
            executor.submit(zpracuj_soubor, cesta, df_goods): cesta
            for cesta in vstupni_soubory
        }
        for future in as_completed(future_to_file):
            zprava, df = future.result()
            print(zprava)
            if df is not None:
                vysledky.append(df)

    # --- Výstup ---
    if vysledky:
        final_df = pd.concat(vysledky, ignore_index=True)
        vystup = os.path.join(slozka, "doporučení_aktivace_EBC.xlsx")
        final_df.to_excel(vystup, index=False)
        print(f"\n✅ HOTOVO! Výstup uložen jako:\n{vystup}")
    else:
        print("⚠️ Nebyly nalezeny žádné položky pro zpracování.")

    input("\nStiskněte Enter pro ukončení...")


if __name__ == "__main__":
    multiprocessing.freeze_support()
    main()

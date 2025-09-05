
# -*- coding: utf-8 -*-
import os
import sys
import traceback
import multiprocessing
from concurrent.futures import ProcessPoolExecutor, as_completed
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def zpracuj_soubor(cesta: str, df_goods: pd.DataFrame):
    """Zpracuje jeden vstupní .xlsx soubor se seznamem položek a EBC sloupci."""
    filename = os.path.basename(cesta)
    print(f"\n🔄 Zpracovávám soubor: {filename}")

    # ⬇️ přeskočit lock/temporary soubory Excelu
    if filename.startswith("~$"):
        return f"ℹ️ {filename}: přeskočeno (dočasný/lock soubor).", None

    # Bezpečné načtení prvního listu
    try:
        wb_input = load_workbook(cesta, read_only=True)
        sheet_names_input = wb_input.sheetnames
        if not sheet_names_input:
            return f"⚠️ {filename}: žádné listy – pravděpodobně poškozený nebo neplatný soubor.", None
        df = pd.read_excel(cesta, sheet_name=sheet_names_input[0])
        df.columns = df.columns.astype(str).str.strip()
    except (BadZipFile, InvalidFileException):
        return f"ℹ️ {filename}: přeskočeno (neplatný/rozpracovaný .xlsx).", None
    except PermissionError:
        return f"ℹ️ {filename}: přeskočeno (soubor je právě uzamčen).", None
    except Exception as e:
        return f"ℹ️ {filename}: přeskočeno ({e}).", None

    # 1) Identifikátor (první sloupec) a EBC sloupce
    if df.shape[1] == 0:
        return f"⚠️ {filename}: prázdný list.", None

    col_oid_left = df.columns[0]

    # Hledání EBC sloupců podle názvů
    ebc_col = next((c for c in df.columns if "EBC požadovaný stav" in c), None)
    ebc2_col = next((c for c in df.columns if "EBC 2 požadovaný stav" in c), None)

    if not ebc_col or not ebc2_col:
        return f"⚠️ {filename}: chybí sloupce 'EBC požadovaný stav' a/nebo 'EBC 2 požadovaný stav'.", None

    # 2) Filtr položek s EBC = 'ano' (case-insensitive, bezpečně pro NaN)
    df_filtered = df[
        (df[ebc_col].astype(str).str.lower() == "ano") |
        (df[ebc2_col].astype(str).str.lower() == "ano")
    ].copy()

    if df_filtered.empty:
        return f"ℹ️ {filename}: žádné položky s EBC = ano.", None

    # 3) Odstranění kolizních sloupců před mergem
    for col in ("Fotografie", "Anotace", "OID_zbozi"):
        if col in df_filtered.columns:
            df_filtered.drop(columns=col, inplace=True)

    # 4) Merge s goods podle OID
    df_filtered = df_filtered.merge(
        df_goods[["OID_zbozi", "Fotografie", "Anotace"]],
        left_on=col_oid_left,
        right_on="OID_zbozi",
        how="left"
    )

    # 5) Zjištění názvů sloupců po mergi (fallback vytvoření)
    foto_candidates = [c for c in df_filtered.columns if c.strip().lower().startswith("fotografie")]
    anotace_candidates = [c for c in df_filtered.columns if c.strip().lower().startswith("anotace")]

    if not foto_candidates:
        if "Fotografie" not in df_filtered.columns:
            df_filtered["Fotografie"] = pd.NA
        foto_col = "Fotografie"
    else:
        foto_col = foto_candidates[-1]

    if not anotace_candidates:
        if "Anotace" not in df_filtered.columns:
            df_filtered["Anotace"] = pd.NA
        anotace_col = "Anotace"
    else:
        anotace_col = anotace_candidates[-1]

    # 6) Vyhodnocení stavu dat (fotky/anotace)
    def check_completeness(row):
        has_photo = pd.notna(row[foto_col]) and str(row[foto_col]).strip() != ""
        has_annotation = pd.notna(row[anotace_col]) and str(row[anotace_col]).strip() != ""
        if has_photo and has_annotation:
            return "Fotografie + Anotace"
        elif has_photo:
            return "Pouze fotografie"
        elif has_annotation:
            return "Pouze anotace"
        else:
            return "Chybí obojí"

    df_filtered["Stav dat"] = df_filtered.apply(check_completeness, axis=1)
    df_filtered["Zdrojový soubor"] = filename

    return f"✅ Zpracováno: {filename}", df_filtered


def main():
    # Složka skriptu
    if getattr(sys, "frozen", False):
        slozka = os.path.dirname(sys.executable)
    else:
        slozka = os.path.dirname(os.path.abspath(__file__))

    # --- Načtení good.xlsx ---
    soubor_good = os.path.join(slozka, "good.xlsx")
    try:
        wb = load_workbook(soubor_good, read_only=True)
        sheet_names = wb.sheetnames
        print(f"✅ Nalezené listy v good.xlsx: {sheet_names}")
        if not sheet_names:
            raise ValueError("Soubor good.xlsx neobsahuje žádné načitatelné listy.")

        df_goods = pd.read_excel(soubor_good, sheet_name=sheet_names[0])
        df_goods.columns = df_goods.columns.astype(str).str.strip()

        # Povinné sloupce: OID_zbozi (kritický), Fotografie/Anotace (dopočítáme, když chybí)
        if "OID_zbozi" not in df_goods.columns:
            raise ValueError("Soubor good.xlsx neobsahuje sloupec 'OID_zbozi' – nelze párovat.")

        if "Fotografie" not in df_goods.columns:
            df_goods["Fotografie"] = pd.NA
        if "Anotace" not in df_goods.columns:
            df_goods["Anotace"] = pd.NA

    except PermissionError:
        print("❌ Chyba: good.xlsx je uzamčen (otevřen v Excelu). Zavřete ho a spusťte znovu.")
        input("\nStiskněte Enter pro ukončení...")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Chyba při načítání good.xlsx: {e}")
        traceback.print_exc()
        input("\nStiskněte Enter pro ukončení...")
        sys.exit(1)

    # --- Sběr vstupních souborů ---
    vstupni_soubory = [
        os.path.join(slozka, f)
        for f in os.listdir(slozka)
        if f.endswith(".xlsx")
        and f.lower() != "good.xlsx"
        and not f.startswith("~$")  # přeskoč lock soubory
    ]

    vysledky = []
    logy = []

    with ProcessPoolExecutor(max_workers=max(1, multiprocessing.cpu_count())) as executor:
        future_to_file = {executor.submit(zpracuj_soubor, cesta, df_goods): cesta for cesta in vstupni_soubory}
        for future in as_completed(future_to_file):
            zprava, df = future.result()
            print(zprava)
            logy.append(zprava)
            if df is not None and not df.empty:
                vysledky.append(df)

    # --- Výstup ---
    if vysledky:
        final_df = pd.concat(vysledky, ignore_index=True)

        # volitelný přehled (souhrn stavů)
        souhrn = final_df["Stav dat"].value_counts(dropna=False).to_dict()
        if souhrn:
            print("\n📊 Souhrn stavů (všechny soubory):")
            for k, v in souhrn.items():
                print(f"  - {k}: {v}")

        vystup = os.path.join(slozka, "doporučení_aktivace_EBC.xlsx")
        try:
            final_df.to_excel(vystup, index=False)
            print(f"\n✅ HOTOVO! Výstup uložen jako:\n{vystup}")
        except PermissionError:
            print(f"❌ Nelze zapsat výstup – {vystup} je otevřený. Zavřete soubor a spusťte znovu.")
    else:
        print("⚠️ Nebyly nalezeny žádné položky pro zpracování.")

    input("\nStiskněte Enter pro ukončení...")


if __name__ == "__main__":
    multiprocessing.freeze_support()
    main()
